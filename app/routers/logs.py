"""
Router para endpoints de logs (inbound).
"""
import datetime
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, status, Request
from fastapi.responses import JSONResponse, Response
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from app.core.db import get_db
from app.models.schemas import LogEntry
from app.services import db_logs, csv_handler
from app.services.slotting_service import slotting_service
from app.services.ai_slotting import ai_slotting
from app.utils.auth import login_required, permission_required, api_login_required
from app.utils.country import get_current_country
from app.core.config import ASYNC_DB_URL
from sqlalchemy.ext.asyncio import create_async_engine
from sqlalchemy import text
import numpy as np
from typing import Optional

# Se mantiene el engine solo para pandas read_sql que requiere una conexión/engine
async_engine = create_async_engine(
    ASYNC_DB_URL,
    pool_pre_ping=True,
    pool_recycle=280,
)

router = APIRouter(prefix="/api", tags=["logs"])


@router.get('/find_item/{item_code}/{import_reference}')
async def find_item(
    request: Request,
    item_code: str, 
    import_reference: str, 
    username: str = Depends(api_login_required), 
    db: AsyncSession = Depends(get_db)
):
    """Busca un item en el maestro y calcula cantidades con sugerencia de Slotting e IA (Multicountry)."""
    item_code = item_code.strip().upper()
    country = get_current_country(request) or "CL"
    
    item_details = await csv_handler.get_item_details_from_master_csv(item_code, country_code=country)
    if item_details is None:
        raise HTTPException(status_code=404, detail=f"Artículo {item_code} no encontrado en el maestro.")
    
    expected_quantity = await csv_handler.get_total_expected_quantity_for_item(item_code, country_code=country)
    original_bin = item_details.get('Bin_1', 'N/A')
    latest_relocated_bin = await db_logs.get_latest_relocated_bin_async(db, item_code, country_code=country)
    effective_bin_location = latest_relocated_bin if latest_relocated_bin else original_bin
    
    # 1. Sugerencia de Slotting Dinámico (Algoritmo Tradicional)
    traditional_suggested_bin = await slotting_service.get_suggested_bin(db, country, item_details)

    # 2. Sugerencia de IA (Aprendizaje Histórico)
    ai_predicted_bin = ai_slotting.predict_best_bin(
        country_code=country,
        item_code=item_code,
        sic_code=item_details.get('SIC_Code_stockroom'),
        fallback_bin=traditional_suggested_bin
    )

    # 3. Validación de Capacidad para la IA
    final_suggested_bin = ai_predicted_bin
    is_ai_prediction = ai_predicted_bin != traditional_suggested_bin

    if is_ai_prediction:
        occupancy = await slotting_service._get_bins_occupancy(db, country)
        current_skus = occupancy.get(ai_predicted_bin.upper(), 0)
        if current_skus >= 4:
            final_suggested_bin = traditional_suggested_bin
            is_ai_prediction = False

    if latest_relocated_bin or final_suggested_bin == effective_bin_location:
        final_suggested_bin = None
        is_ai_prediction = False

    response_data = {
        "itemCode": item_details.get('Item_Code', item_code),
        "description": item_details.get('Item_Description', 'N/A'),
        "binLocation": effective_bin_location,
        "suggestedBin": final_suggested_bin,
        "is_ai_prediction": is_ai_prediction,
        "aditionalBins": item_details.get('Aditional_Bin_Location', 'N/A'),
        "physicalQty": item_details.get('Physical_Qty', '0'),
        "weight": item_details.get('Weight_per_Unit', 'N/A'),
        "defaultQtyGrn": expected_quantity,
        "itemType": item_details.get('ABC_Code_stockroom', 'N/A'),
        "sicCode": item_details.get('SIC_Code_stockroom', 'N/A'),
        "dateLastReceived": item_details.get('Date_Last_Received', 'N/A'),
        "supersededBy": item_details.get('SupersededBy', 'N/A')
    }
    return JSONResponse(content=response_data)


@router.post('/add_log')
async def add_log(request: Request, data: LogEntry, username: str = Depends(permission_required("inbound")), db: AsyncSession = Depends(get_db)):
    """Añade un registro de entrada y entrena la IA de Slotting."""
    if data.quantity <= 0:
        raise HTTPException(status_code=400, detail="Cantidad debe ser > 0")
    
    item_code_form = data.itemCode.strip().upper()
    import_reference = data.importReference
    quantity_received_form = data.quantity
    country = get_current_country(request) or "CL"
    
    item_details = await csv_handler.get_item_details_from_master_csv(item_code_form, country_code=country)
    if item_details is None:
        raise HTTPException(status_code=404, detail=f"Artículo {item_code_form} no encontrado.")
    
    # Obtener la ubicación efectiva
    original_bin_from_master = item_details.get('Bin_1', 'N/A')
    latest_relocated_bin_for_item = await db_logs.get_latest_relocated_bin_async(db, item_code_form, country_code=country)
    bin_to_log_as_original = latest_relocated_bin_for_item if latest_relocated_bin_for_item else original_bin_from_master
    
    total_received_before = await db_logs.get_total_received_for_import_reference_async(db, import_reference, item_code_form, country_code=country)
    total_expected = await csv_handler.get_total_expected_quantity_for_item(item_code_form, country_code=country)
    total_received_now = total_received_before + quantity_received_form
    difference = total_received_now - total_expected
    
    # Usar hora de Colombia (UTC-5)
    colombia_tz = datetime.timezone(datetime.timedelta(hours=-5))
    current_time = datetime.datetime.now(colombia_tz)

    entry_data = {
        'timestamp': current_time.isoformat(timespec='seconds'),
        'importReference': import_reference.strip() if import_reference else '',
        'waybill': data.waybill.strip() if data.waybill else '',
        'itemCode': item_code_form,
        'itemDescription': item_details.get('Item_Description', 'N/A'),
        'binLocation': bin_to_log_as_original,
        'relocatedBin': data.relocatedBin or '',
        'qtyReceived': quantity_received_form,
        'qtyGrn': total_expected,
        'difference': difference,
        'country_code': country
    }

    # APRENDIZAJE: Si el operario eligió una ubicación de reubicación, alimentamos la IA
    if data.relocatedBin:
        ai_slotting.learn_from_decision(
            country_code=country,
            item_code=item_code_form,
            final_bin=data.relocatedBin,
            sic_code=item_details.get('SIC_Code_stockroom')
        )
    
    log_id = await db_logs.save_log_entry_db_async(db, entry_data)
    if log_id:
        log_entry_data_for_response = {"id": log_id, **entry_data}
        return JSONResponse({'message': 'Registro añadido con éxito.', 'entry': log_entry_data_for_response}, status_code=201)
    raise HTTPException(status_code=500, detail="Error al guardar el registro.")


@router.put('/update_log/{log_id}')
async def update_log(request: Request, log_id: int, data: dict, username: str = Depends(permission_required("inbound")), db: AsyncSession = Depends(get_db)):
    """Actualiza un registro de entrada existente."""
    country = get_current_country(request) or "CL"
    existing_log = await db_logs.get_log_entry_by_id_async(db, log_id, country_code=country)
    if not existing_log:
        raise HTTPException(status_code=404, detail=f"Registro con ID {log_id} no encontrado.")
    
    waybill = data.get('waybill', existing_log.get('waybill'))
    relocated_bin = data.get('relocatedBin', existing_log.get('relocatedBin'))
    qty_received = int(data.get('qtyReceived', existing_log.get('qtyReceived')))
    
    import_reference = existing_log['importReference']
    item_code = existing_log['itemCode']
    
    # Recalcular diferencia
    total_received_others = await db_logs.get_total_received_for_import_reference_async(db, import_reference, item_code, country_code=country)
    total_received_others -= int(existing_log.get('qtyReceived', 0))
    total_received_now = total_received_others + qty_received
    total_expected = await csv_handler.get_total_expected_quantity_for_item(item_code, country_code=country)
    difference = total_received_now - total_expected
    
    entry_data_for_db = {
        'waybill': waybill.strip() if waybill else '',
        'relocatedBin': relocated_bin.strip() if relocated_bin else '',
        'qtyReceived': qty_received,
        'difference': difference,
        'timestamp': datetime.datetime.now().isoformat(timespec='seconds')
    }
    
    success = await db_logs.update_log_entry_db_async(db, log_id, entry_data_for_db, country_code=country)
    if success:
        return JSONResponse({'message': f'Registro {log_id} actualizado con éxito.'})
    raise HTTPException(status_code=500, detail="Error al actualizar el registro.")


@router.get('/get_logs')
async def get_logs(request: Request, version_date: Optional[str] = None, username: str = Depends(permission_required("inbound")), db: AsyncSession = Depends(get_db)):
    """
    Obtiene los registros de entrada.
    """
    country = get_current_country(request) or "CL"
    if version_date:
        logs = await db_logs.load_archived_log_data_db_async(db, country, version_date)
    else:
        logs = await db_logs.load_log_data_db_async(db, country)
    return JSONResponse(content=logs)


@router.post('/logs/archive')
async def archive_logs(request: Request, username: str = Depends(permission_required("inbound")), db: AsyncSession = Depends(get_db)):
    """Archiva los logs actuales para limpiar la base."""
    country = get_current_country(request) or "CL"
    success = await db_logs.archive_current_logs_db_async(db, country_code=country)
    if success:
        return JSONResponse({'message': 'Logs archivados correctamente.'})
    raise HTTPException(status_code=500, detail="Error al archivar los logs.")


@router.get('/logs/versions')
async def get_log_versions(request: Request, username: str = Depends(permission_required("inbound")), db: AsyncSession = Depends(get_db)):
    """Obtiene las fechas disponibles de archivos históricos."""
    country = get_current_country(request) or "CL"
    versions = await db_logs.get_archived_versions_db_async(db, country_code=country)
    return JSONResponse(content=versions)


@router.delete('/delete_log/{log_id}')
async def delete_log_api(request: Request, log_id: int, username: str = Depends(permission_required("inbound")), db: AsyncSession = Depends(get_db)):
    """Elimina un registro de entrada."""
    country = get_current_country(request) or "CL"
    success = await db_logs.delete_log_entry_db_async(db, log_id, country_code=country)
    if success:
        return JSONResponse({'message': f'Registro {log_id} eliminado con éxito.'})
    raise HTTPException(status_code=404, detail=f"Registro con ID {log_id} no encontrado.")


@router.get('/export_log')
async def export_log(request: Request, version_date: Optional[str] = None, username: str = Depends(permission_required("inbound")), db: AsyncSession = Depends(get_db)):
    """Exporta todos los registros de inbound a un archivo Excel."""
    country = get_current_country(request) or "CL"
    if version_date:
        logs_data = await db_logs.load_archived_log_data_db_async(db, country, version_date)
    else:
        logs_data = await db_logs.load_log_data_db_async(db, country)
    
    if not logs_data:
        raise HTTPException(status_code=404, detail="No hay registros para exportar")

    df = pd.DataFrame(logs_data)

    # Procesar timestamp para asegurar hora local correcta en Excel
    try:
        # Convertir a datetime
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        
        # Si tiene zona horaria (los nuevos registros), convertir a Colombia y quitar tz info para que Excel lo muestre limpio
        if df['timestamp'].dt.tz is not None:
             colombia_tz = datetime.timezone(datetime.timedelta(hours=-5))
             df['timestamp'] = df['timestamp'].dt.tz_convert(colombia_tz).dt.tz_localize(None)
    except Exception as e:
        print(f"Advertencia procesando fechas en export: {e}")

    df_export = df[[
        'timestamp', 'importReference', 'waybill', 'itemCode', 'itemDescription',
        'binLocation', 'relocatedBin', 'qtyReceived', 'qtyGrn', 'difference'
    ]].rename(columns={
        'timestamp': 'Timestamp', 'importReference': 'Import Reference', 'waybill': 'Waybill',
        'itemCode': 'Item Code', 'itemDescription': 'Item Description',
        'binLocation': 'Bin Location (Original)', 'relocatedBin': 'Relocated Bin (New)',
        'qtyReceived': 'Qty. Received', 'qtyGrn': 'Qty. Expected (Total)', 'difference': 'Difference'
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='InboundLogCompleto')
        worksheet = writer.sheets['InboundLogCompleto']
        for i, col_name in enumerate(df_export.columns):
            column_letter = get_column_letter(i + 1)
            max_len = max(df_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[column_letter].width = max_len

    output.seek(0)
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inbound_log_completo_{timestamp_str}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get('/items_without_grn')
async def get_items_without_grn(request: Request, username: str = Depends(permission_required("inbound"))):
    """Obtiene un reporte de items en el log que no están en ningún GRN (Multicountry)."""
    country = get_current_country(request) or "CL"
    try:
        async with async_engine.connect() as conn:
            query = text('SELECT * FROM logs WHERE country_code = :country')
            logs_df = await conn.run_sync(lambda sync_conn: pd.read_sql_query(query, sync_conn, params={"country": country}))

        grn_df = csv_handler.df_grn_cache.get(country)

        if logs_df.empty:
            return JSONResponse(content={"data": [], "message": "No hay registros en el log"})

        # Convertir qtyReceived a numérico
        logs_df['qtyReceived'] = pd.to_numeric(logs_df['qtyReceived'], errors='coerce').fillna(0)

        if grn_df is None or grn_df.empty:
            # Si no hay datos de GRN, todos los items del log están "sin GRN"
            items_without_grn = logs_df.groupby('itemCode').agg({
                'itemDescription': 'first',
                'importReference': 'first',
                'waybill': 'first',
                'qtyReceived': 'sum'
            }).reset_index()
        else:
            # Obtener items únicos en el log
            log_items = set(logs_df['itemCode'].unique())
            # Obtener items únicos en el GRN
            grn_items = set(grn_df['Item_Code'].unique())
            # Items que están en el log pero NO en el GRN
            items_not_in_grn = log_items - grn_items

            # Filtrar el log para obtener solo esos items
            filtered_logs = logs_df[logs_df['itemCode'].isin(items_not_in_grn)]
            # Agrupar por item y sumar cantidades
            items_without_grn = filtered_logs.groupby('itemCode').agg({
                'itemDescription': 'first',
                'importReference': 'first',
                'waybill': 'first',
                'qtyReceived': 'sum'
            }).reset_index()

        if items_without_grn.empty:
            return JSONResponse(content={"data": [], "message": "Todos los items están asociados con GRNs"})

        # Ordenar por código de item
        items_without_grn = items_without_grn.sort_values('itemCode').reset_index(drop=True)

        # Convertir cantidad a entero
        items_without_grn['qtyReceived'] = items_without_grn['qtyReceived'].astype(int)

        # Renombrar columnas
        items_without_grn = items_without_grn.rename(columns={
            'importReference': 'Import Ref.',
            'waybill': 'Waybill',
            'itemCode': 'Código de Ítem',
            'itemDescription': 'Descripción',
            'qtyReceived': 'Cantidad Recibida'
        })

        # Reordenar columnas en el orden especificado
        items_without_grn = items_without_grn[['Import Ref.', 'Waybill', 'Código de Ítem', 'Descripción', 'Cantidad Recibida']]

        return JSONResponse(content={
            "data": items_without_grn.to_dict(orient='records'),
            "columns": items_without_grn.columns.tolist(),
            "message": f"Se encontraron {len(items_without_grn)} items sin asociación con GRN"
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get('/export_items_without_grn')
async def export_items_without_grn(request: Request, timezone_offset: int = 0, username: str = Depends(permission_required("inbound"))):
    """Exporta el reporte de items sin GRN a Excel (Multicountry)."""
    country = get_current_country(request) or "CL"
    try:
        async with async_engine.connect() as conn:
            query = text('SELECT * FROM logs WHERE country_code = :country')
            logs_df = await conn.run_sync(lambda sync_conn: pd.read_sql_query(query, sync_conn, params={"country": country}))

        grn_df = csv_handler.df_grn_cache.get(country)

        if logs_df.empty:
            raise HTTPException(status_code=404, detail="No hay registros en el log")

        # Convertir qtyReceived a numérico
        logs_df['qtyReceived'] = pd.to_numeric(logs_df['qtyReceived'], errors='coerce').fillna(0)

        if grn_df is None or grn_df.empty:
            items_without_grn = logs_df.groupby('itemCode').agg({
                'itemDescription': 'first',
                'importReference': 'first',
                'waybill': 'first',
                'qtyReceived': 'sum'
            }).reset_index()
        else:
            log_items = set(logs_df['itemCode'].unique())
            grn_items = set(grn_df['Item_Code'].unique())
            items_not_in_grn = log_items - grn_items
            filtered_logs = logs_df[logs_df['itemCode'].isin(items_not_in_grn)]
            items_without_grn = filtered_logs.groupby('itemCode').agg({
                'itemDescription': 'first',
                'importReference': 'first',
                'waybill': 'first',
                'qtyReceived': 'sum'
            }).reset_index()

        if items_without_grn.empty:
            raise HTTPException(status_code=404, detail="Todos los items están asociados con GRNs")

        items_without_grn = items_without_grn.sort_values('itemCode').reset_index(drop=True)

        # Convertir cantidad a entero
        items_without_grn['qtyReceived'] = items_without_grn['qtyReceived'].astype(int)

        df_for_export = items_without_grn.rename(columns={
            'itemCode': 'Código de Ítem',
            'itemDescription': 'Descripción',
            'importReference': 'Import Ref.',
            'waybill': 'Waybill',
            'qtyReceived': 'Cantidad Recibida'
        })

        # Reordenar columnas en el orden especificado
        df_for_export = df_for_export[['Import Ref.', 'Waybill', 'Código de Ítem', 'Descripción', 'Cantidad Recibida']]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_for_export.to_excel(writer, index=False, sheet_name='Items Sin GRN')
            worksheet = writer.sheets['Items Sin GRN']
            for i, col_name in enumerate(df_for_export.columns):
                column_letter = get_column_letter(i + 1)
                max_len = max(df_for_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[column_letter].width = max_len

        output.seek(0)
        # Calcular fecha/hora del cliente usando el offset (minutos)
        utc_now = datetime.datetime.now(datetime.timezone.utc)
        client_time = utc_now - datetime.timedelta(minutes=timezone_offset)
        timestamp_str = client_time.strftime("%Y%m%d_%H%M%S")
        
        filename = f"items_sin_grn_{timestamp_str}.xlsx"
        return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename={filename}"})

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar: {str(e)}")


@router.get('/export_reconciliation')
async def export_reconciliation(request: Request, timezone_offset: int = 0, archive_date: Optional[str] = None, username: str = Depends(permission_required("inbound")), db: AsyncSession = Depends(get_db)):
    """Genera y exporta el reporte de conciliación (Multicountry)."""
    country = get_current_country(request) or "CL"
    try:
        async with async_engine.connect() as conn:
            if archive_date:
                query = text('SELECT * FROM logs WHERE archived_at = :date AND country_code = :country')
                logs_df = await conn.run_sync(lambda sync_conn: pd.read_sql_query(query, sync_conn, params={"date": archive_date, "country": country}))
            else:
                query = text('SELECT * FROM logs WHERE archived_at IS NULL AND country_code = :country')
                logs_df = await conn.run_sync(lambda sync_conn: pd.read_sql_query(query, sync_conn, params={"country": country}))

        # Accedemos al caché de GRN del país
        grn_df = csv_handler.df_grn_cache.get(country)

        if logs_df.empty or grn_df is None:
            raise HTTPException(status_code=404, detail="No hay datos suficientes para generar la conciliación")

        logs_df['qtyReceived'] = pd.to_numeric(logs_df['qtyReceived'], errors='coerce').fillna(0)
        grn_df['Quantity'] = pd.to_numeric(grn_df['Quantity'], errors='coerce').fillna(0)

        # Calcular totales recibidos por ítem desde el log
        item_totals = logs_df.groupby(['itemCode'])['qtyReceived'].sum().reset_index()
        item_totals = item_totals.rename(columns={'itemCode': 'Item_Code', 'qtyReceived': 'Total_Recibido'})

        # Calcular totales esperados por ítem (sumando todas las líneas del GRN para ese ítem)
        item_expected_totals = grn_df.groupby(['Item_Code'])['Quantity'].sum().reset_index()
        item_expected_totals = item_expected_totals.rename(columns={'Quantity': 'Total_Esperado_Item'})

        # NO agrupar el GRN - mantener todas las líneas individuales
        grn_lines = grn_df[['GRN_Number', 'Item_Code', 'Item_Description', 'Quantity']].copy()
        grn_lines = grn_lines.rename(columns={'Quantity': 'Cant_Esperada_Linea'})

        # Combinar cada línea del GRN con los totales recibidos y esperados del ítem
        merged_df = pd.merge(grn_lines, item_totals, on='Item_Code', how='left')
        merged_df = pd.merge(merged_df, item_expected_totals, on='Item_Code', how='left')

        if not logs_df.empty:
            logs_df['id'] = pd.to_numeric(logs_df['id'])
            latest_logs = logs_df.sort_values('id', ascending=False).drop_duplicates('itemCode')
            
            # Extraer tanto binLocation como relocatedBin por separado
            locations_df = latest_logs[['itemCode', 'binLocation', 'relocatedBin']].rename(
                columns={'itemCode': 'Item_Code', 'binLocation': 'Bin_Original', 'relocatedBin': 'Bin_Reubicado'}
            )
            
            merged_df = pd.merge(merged_df, locations_df, on='Item_Code', how='left')

        merged_df['Total_Recibido'] = merged_df['Total_Recibido'].fillna(0)
        merged_df['Cant_Esperada_Linea'] = merged_df['Cant_Esperada_Linea'].fillna(0)
        merged_df['Total_Esperado_Item'] = merged_df['Total_Esperado_Item'].fillna(0)
        merged_df['Diferencia'] = merged_df['Total_Recibido'] - merged_df['Total_Esperado_Item']

        merged_df.fillna({'Bin_Original': 'N/A', 'Bin_Reubicado': ''}, inplace=True)

        merged_df['Total_Recibido'] = merged_df['Total_Recibido'].astype(int)
        merged_df['Cant_Esperada_Linea'] = merged_df['Cant_Esperada_Linea'].astype(int)
        merged_df['Total_Esperado_Item'] = merged_df['Total_Esperado_Item'].astype(int)
        merged_df['Diferencia'] = merged_df['Diferencia'].astype(int)

        # Ordenar por GRN ascendente
        merged_df = merged_df.sort_values('GRN_Number', ascending=True)

        df_for_export = merged_df.rename(columns={
            'GRN_Number': 'GRN',
            'Item_Code': 'Código de Ítem',
            'Item_Description': 'Descripción',
            'Bin_Original': 'Ubicación',
            'Bin_Reubicado': 'Reubicado',
            'Cant_Esperada_Linea': 'Cant. Esperada',
            'Total_Recibido': 'Cant. Recibida',
            'Diferencia': 'Diferencia'
        })
        
        cols_order = ['GRN', 'Código de Ítem', 'Descripción', 'Ubicación', 'Reubicado', 'Cant. Esperada', 'Cant. Recibida', 'Diferencia']
        df_for_export = df_for_export[cols_order]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_for_export.to_excel(writer, index=False, sheet_name='ReporteDeConciliacion')
            worksheet = writer.sheets['ReporteDeConciliacion']
            for i, col_name in enumerate(df_for_export.columns):
                column_letter = get_column_letter(i + 1)
                max_len = max(df_for_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[column_letter].width = max_len

        output.seek(0)
        # Calcular fecha/hora del cliente usando el offset (minutos)
        utc_now = datetime.datetime.now(datetime.timezone.utc)
        client_time = utc_now - datetime.timedelta(minutes=timezone_offset)
        timestamp_str = client_time.strftime("%Y%m%d_%H%M%S")
        
        filename = f"reporte_conciliacion_{timestamp_str}.xlsx"
        return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename={filename}"})

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error interno al generar el archivo de conciliación: {e}")
