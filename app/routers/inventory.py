"""
Router para endpoints de gestión de inventario y conteos administrativos.
"""
import datetime
import pandas as pd
from io import BytesIO
from urllib.parse import urlencode
from typing import Optional, Dict, Any
import numpy as np
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Request, Depends, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy import select, func, delete, insert, update, text

from app.core.config import ASYNC_DB_URL
from app.core.db import get_db
from app.core.templates import templates
from app.services.csv_handler import master_qty_map
from app.services import db_counts
from app.utils.auth import login_required, admin_login_required, permission_required
from app.utils.country import get_current_country
from app.models.sql_models import AppState, StockCount, CountSession, RecountList, SessionLocation, MasterItem
from app.services.csv_to_db import sync_master_csv_to_db

# --- Inicialización ---
router = APIRouter(tags=["inventory"])
async_engine = create_async_engine(ASYNC_DB_URL)


async def get_inventory_summary_stats(db: AsyncSession, country_code: str) -> Optional[Dict[str, Any]]:
    """Calcula y devuelve un resumen de estadísticas para el panel de admin de inventario del país."""
    summary = {
        'general': {
            'total_items_master': 0,
        },
        'stages': {}
    }
    
    try:
        # --- Estadísticas Generales (del maestro de items) ---
        country_master_qty = master_qty_map.get(country_code, {})
        if country_master_qty:
            total_items_with_stock = sum(1 for qty in country_master_qty.values() if qty is not None and qty > 0)
            summary['general']['total_items_master'] = total_items_with_stock

        # --- Estadísticas por Etapa ---
        for stage_num in range(1, 5):
            # Items contados en esta etapa
            stmt_items_counted = select(func.count(func.distinct(StockCount.item_code))).\
                join(CountSession, StockCount.session_id == CountSession.id).\
                where(CountSession.inventory_stage == stage_num, CountSession.country_code == country_code, StockCount.country_code == country_code)
            
            items_counted = (await db.execute(stmt_items_counted)).scalar() or 0

            # Si no se contó nada en esta etapa, podemos saltarla
            if items_counted == 0:
                continue

            # Total de unidades contadas
            stmt_total_units = select(func.sum(StockCount.counted_qty)).\
                join(CountSession, StockCount.session_id == CountSession.id).\
                where(CountSession.inventory_stage == stage_num, CountSession.country_code == country_code, StockCount.country_code == country_code)
            
            total_units_counted = (await db.execute(stmt_total_units)).scalar() or 0
            
            # Calcular diferencias para esta etapa
            stmt_diff = select(StockCount.item_code, func.sum(StockCount.counted_qty).label('total_counted')).\
                join(CountSession, StockCount.session_id == CountSession.id).\
                where(CountSession.inventory_stage == stage_num, CountSession.country_code == country_code, StockCount.country_code == country_code).\
                group_by(StockCount.item_code)
            
            counted_items_result = (await db.execute(stmt_diff)).all()
            counted_items_map = {row.item_code: row.total_counted for row in counted_items_result}

            items_with_discrepancy = 0
            for item_code, total_counted in counted_items_map.items():
                system_qty_raw = country_master_qty.get(item_code)
                system_qty = 0
                if system_qty_raw is not None:
                    try:
                        system_qty = int(float(system_qty_raw))
                    except (ValueError, TypeError):
                        system_qty = 0
                
                if total_counted != system_qty:
                    items_with_discrepancy += 1
            
            # Precisión del conteo
            accuracy = 0
            if items_counted > 0:
                accuracy = ((items_counted - items_with_discrepancy) / items_counted) * 100
            
            # Efectividad de Cobertura
            coverage_effectiveness = 0
            total_items_master_with_stock = summary['general'].get('total_items_master', 0)
            if total_items_master_with_stock > 0:
                items_correctly_counted = items_counted - items_with_discrepancy
                coverage_effectiveness = (items_correctly_counted / total_items_master_with_stock) * 100

            # Guardar estadísticas de la etapa
            summary['stages'][stage_num] = {
                'items_counted': items_counted,
                'total_units_counted': total_units_counted,
                'items_with_discrepancy': items_with_discrepancy,
                'accuracy': f"{accuracy:.2f}%",
                'coverage_effectiveness': f"{coverage_effectiveness:.2f}%"
            }

        # --- Items en lista de reconteo (para etapas futuras) ---
        for stage_to_check in range(2, 5):
            stmt_recount = select(func.count(RecountList.item_code)).where(RecountList.stage_to_count == stage_to_check, RecountList.country_code == country_code)
            items_in_recount_list = (await db.execute(stmt_recount)).scalar() or 0
            
            if stage_to_check in summary['stages']:
                summary['stages'][stage_to_check]['items_in_recount_list'] = items_in_recount_list
            elif items_in_recount_list > 0:
                 # Si la etapa aún no tiene conteos pero ya hay lista de reconteo
                summary['stages'][stage_to_check] = { 'items_in_recount_list': items_in_recount_list }

    except Exception as e:
        print(f"Error al calcular estadísticas de inventario para {country_code}: {e}")
        return None

    return summary


# ===== RUTAS DE ADMIN INVENTORY =====

@router.get('/admin_inventory', response_class=RedirectResponse)
async def redirect_admin_inventory():
    """Redirección legacy."""
    return RedirectResponse(url='/admin/inventory')


@router.get('/admin/inventory', response_class=HTMLResponse, name='admin_inventory')
async def admin_inventory_get(request: Request, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Página principal de administración de inventario."""
    country = get_current_country(request) or "CL"
    
    result = await db.execute(select(AppState).where(AppState.key == 'current_inventory_stage', AppState.country_code == country))
    stage = result.scalar_one_or_none()
    
    if not stage:
        # Si no existe, inicializamos a etapa 0 (inactivo)
        new_stage = AppState(key='current_inventory_stage', value='0', country_code=country)
        db.add(new_stage)
        await db.commit()
        await db.refresh(new_stage)
        stage = new_stage

    message = request.query_params.get('message')
    error = request.query_params.get('error')
    
    summary_stats = await get_inventory_summary_stats(db, country)

    return templates.TemplateResponse('admin_inventory.html', {
        "request": request, 
        "stage": stage,
        "message": message,
        "error": error,
        "summary": summary_stats
    })


@router.post('/admin/inventory/start_stage_1', name='start_inventory_stage_1')
async def start_inventory_stage_1(request: Request, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Inicia un nuevo ciclo de inventario en Etapa 1."""
    
    try:
        country = get_current_country(request) or "CL"
        print(f"Limpiando tablas de inventario para un nuevo ciclo en {country}...")
        await db.execute(delete(StockCount).where(StockCount.country_code == country))
        await db.execute(delete(CountSession).where(CountSession.country_code == country))
        await db.execute(delete(SessionLocation).where(SessionLocation.country_code == country))
        await db.execute(delete(RecountList).where(RecountList.country_code == country))
        
        print(f"Tablas de inventario limpiadas para {country}.")

        # Sincronizar maestro de items desde CSV a DB
        print(f"Sincronizando maestro de items para {country}...")
        await sync_master_csv_to_db(db, country_code=country)
        print(f"Sincronización completada para {country}.")

        # Actualizar estado
        stmt_update = update(AppState).where(AppState.key == 'current_inventory_stage', AppState.country_code == country).values(value='1')
        await db.execute(stmt_update)
        
        await db.commit()
        
        query_params = urlencode({"message": f"Inventario reiniciado en Etapa 1 para {country}. Todos los datos han sido reseteados."})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)
    except Exception as e:
        query_params = urlencode({"error": f"Error de base de datos: {e}"})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)


@router.post('/admin/inventory/advance/{next_stage}', name='advance_inventory_stage')
async def advance_inventory_stage(request: Request, next_stage: int, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Avanza el inventario a la siguiente etapa."""
    country = get_current_country(request) or "CL"
    prev_stage = next_stage - 1
    
    try:
        # Calcular items contados en etapa previa
        stmt = select(StockCount.item_code, func.sum(StockCount.counted_qty).label('total_counted')).\
            join(CountSession, StockCount.session_id == CountSession.id).\
            where(CountSession.inventory_stage == prev_stage, CountSession.country_code == country, StockCount.country_code == country).\
            group_by(StockCount.item_code)
        
        result = await db.execute(stmt)
        counted_items = result.all()
        
        # Limpiar lista de reconteo anterior para esta etapa
        await db.execute(delete(RecountList).where(RecountList.stage_to_count == next_stage, RecountList.country_code == country))

        items_for_recount = []
        country_master_qty = master_qty_map.get(country, {})
        for item in counted_items:
            item_code = item.item_code
            total_counted = item.total_counted
            
            system_qty = country_master_qty.get(item_code)
            system_qty = int(system_qty) if system_qty is not None else 0

            if total_counted != system_qty:
                items_for_recount.append({"item_code": item_code, "stage_to_count": next_stage, "country_code": country})

        if items_for_recount:
            await db.execute(insert(RecountList), items_for_recount)

        # Actualizar estado de la aplicación
        stmt_update = update(AppState).where(AppState.key == 'current_inventory_stage', AppState.country_code == country).values(value=str(next_stage))
        await db.execute(stmt_update)
        
        await db.commit()

        message = f"Proceso completado para {country}. Etapa de inventario avanzada a {next_stage}. Se encontraron {len(items_for_recount)} items con diferencias."
        query_params = urlencode({"message": message})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)

    except Exception as e:
        query_params = urlencode({"error": f"Error inesperado: {e}"})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)


@router.post('/admin/inventory/finalize', name='finalize_inventory')
async def finalize_inventory(request: Request, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Finaliza el ciclo de inventario para el país."""
    
    try:
        country = get_current_country(request) or "CL"
        stmt_update = update(AppState).where(AppState.key == 'current_inventory_stage', AppState.country_code == country).values(value='0')
        await db.execute(stmt_update)
        await db.commit()
        
        query_params = urlencode({"message": f"Ciclo de inventario finalizado y cerrado para {country}."})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)
    except Exception as e:
        query_params = urlencode({"error": f"Error de base de datos: {e}"})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)


@router.get('/admin/inventory/report', name='generate_inventory_report')
async def generate_inventory_report(request: Request, user: str = Depends(permission_required("inventory"))):
    """Genera un reporte Excel del inventario."""
    country = get_current_country(request) or "CL"
    try:
        # Usamos pandas read_sql con connection para queries complejos de reporte
        async with async_engine.connect() as conn:
            query = """
                SELECT
                    sc.item_code,
                    sc.item_description,
                    cs.inventory_stage,
                    sc.counted_qty
                FROM stock_counts sc
                JOIN count_sessions cs ON sc.session_id = cs.id
                WHERE sc.country_code = :country AND cs.country_code = :country
            """
            # Pandas read_sql espera una conexión raw o compatible, usamos run_sync
            all_counts_df = await conn.run_sync(lambda sync_conn: pd.read_sql_query(text(query), sync_conn, params={"country": country}))

        if all_counts_df.empty:
            query_params = urlencode({"error": f"No hay datos de conteo para generar un informe en {country}."})
            return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)

        stage_counts = all_counts_df.groupby(['item_code', 'item_description', 'inventory_stage'])['counted_qty'].sum().reset_index()

        pivot_df = stage_counts.pivot_table(
            index=['item_code', 'item_description'],
            columns='inventory_stage',
            values='counted_qty',
            aggfunc='sum'
        ).fillna(0)

        pivot_df.columns = [f'Conteo Etapa {int(col)}' for col in pivot_df.columns]
        
        country_master_qty = master_qty_map.get(country, {})
        system_qtys = pd.Series(country_master_qty, name='Cantidad Sistema').astype('float64').fillna(0)
        
        report_df = pivot_df.join(system_qtys, on='item_code').fillna(0)
        report_df.rename_axis(index={'item_code': 'Item Code', 'item_description': 'Description'}, inplace=True)
        report_df.reset_index(inplace=True)

        final_qty = pd.Series(0, index=report_df.index)
        for stage in sorted([int(c.split()[-1]) for c in pivot_df.columns], reverse=True):
            stage_col = f'Conteo Etapa {stage}'
            final_qty = np.where(final_qty == 0, report_df.get(stage_col, 0), final_qty)
        report_df['Cantidad Final Contada'] = final_qty

        report_df['Diferencia Final'] = report_df['Cantidad Final Contada'] - report_df['Cantidad Sistema']

        cols = ['Item Code', 'Description', 'Cantidad Sistema']
        stage_cols = sorted([col for col in report_df.columns if 'Conteo Etapa' in col])
        final_cols = ['Cantidad Final Contada', 'Diferencia Final']
        report_df = report_df[cols + stage_cols + final_cols]

        for col in report_df.columns:
            if 'Cantidad' in col or 'Conteo' in col or 'Diferencia' in col:
                report_df[col] = report_df[col].astype(int)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='InformeFinalInventario')
            worksheet = writer.sheets['InformeFinalInventario']
            for i, col_name in enumerate(report_df.columns):
                column_letter = get_column_letter(i + 1)
                max_len = max(report_df[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[column_letter].width = max_len
        
        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"informe_final_inventario_{timestamp_str}.xlsx"
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Error generando el informe de inventario: {e}")
        query_params = urlencode({"error": f"No se pudo generar el informe: {str(e)}"})
        return RedirectResponse(url=f"/admin/inventory?{query_params}", status_code=status.HTTP_302_FOUND)


@router.get('/api/export_recount_list/{stage_number}', name='export_recount_list')
async def export_recount_list(request: Request, stage_number: int, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Exporta la lista de items a recontar para una etapa específica."""
    country = get_current_country(request) or "CL"

    result = await db.execute(select(RecountList.item_code).where(RecountList.stage_to_count == stage_number, RecountList.country_code == country))
    items_to_recount = result.all() # list of Row objects

    if not items_to_recount:
        raise HTTPException(status_code=404, detail=f"No hay items en la lista de reconteo para la Etapa {stage_number} en {country}.")

    # Importar la función para obtener detalles del item
    from app.services.csv_handler import get_item_details_from_master_csv
    
    enriched_data = []
    for row in items_to_recount:
        item_code = row.item_code
        details = await get_item_details_from_master_csv(item_code, country_code=country)
        if details:
            enriched_data.append({
                'Código de Item': item_code,
                'Descripción': details.get('Item_Description', 'N/A'),
                'Ubicación en Sistema': details.get('Bin_1', 'N/A')
            })
        else:
            # Para items "fantasma" que no están en el maestro
            enriched_data.append({
                'Código de Item': item_code,
                'Descripción': 'ITEM NO ENCONTRADO EN MAESTRO',
                'Ubicación en Sistema': 'N/A'
            })

    df = pd.DataFrame(enriched_data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'Reconteo_{country}_Et{stage_number}')
        worksheet = writer.sheets[f'Reconteo_{country}_Et{stage_number}']
        for i, col_name in enumerate(df.columns):
            column_letter = get_column_letter(i + 1)
            max_len = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[column_letter].width = max_len
    
    output.seek(0)
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"lista_reconteo_{country}_etapa_{stage_number}_{timestamp_str}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ===== APIs PARA REACT ADMIN INVENTORY =====

@router.get('/api/admin/inventory/summary')
async def get_inventory_summary_api(request: Request, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """API: Obtiene el resumen del estado del inventario."""
    country = get_current_country(request) or "CL"
    stats = await get_inventory_summary_stats(db, country)
    
    # Obtener estado actual
    result = await db.execute(select(AppState).where(AppState.key == 'current_inventory_stage', AppState.country_code == country))
    stage_state = result.scalar_one_or_none()
    current_stage = int(stage_state.value) if stage_state else 0
    
    from fastapi.responses import JSONResponse
    return JSONResponse(content={
        "stage": current_stage,
        "stats": stats
    })

@router.post('/api/admin/inventory/start_stage_1')
async def start_inventory_stage_1_api(request: Request, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """API: Inicia Etapa 1."""
    country = get_current_country(request) or "CL"
    # Reset Current Stage to 1
    result = await db.execute(select(AppState).where(AppState.key == 'current_inventory_stage', AppState.country_code == country))
    stage_state = result.scalar_one_or_none()
    if not stage_state:
        stage_state = AppState(key='current_inventory_stage', value='1', country_code=country)
        db.add(stage_state)
    else:
        stage_state.value = '1'
    
    # Limpiar tablas (logica simplificada de start_inventory_stage_1)
    await db.execute(delete(StockCount).where(StockCount.country_code == country))
    await db.execute(delete(CountSession).where(CountSession.country_code == country))
    await db.execute(delete(SessionLocation).where(SessionLocation.country_code == country))
    await db.execute(delete(RecountList).where(RecountList.country_code == country))
    
    await db.commit()
    from fastapi.responses import JSONResponse
    return JSONResponse(content={"message": f"Inventario Etapa 1 en {country} iniciado correctamente", "stage": 1})

@router.post('/api/admin/inventory/advance_stage/{next_stage}')
async def advance_inventory_stage_api(request: Request, next_stage: int, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """API: Avanza etapa."""
    country = get_current_country(request) or "CL"
    # Validar next_stage logic...
    result = await db.execute(select(AppState).where(AppState.key == 'current_inventory_stage', AppState.country_code == country))
    stage_state = result.scalar_one_or_none()
    current_stage = int(stage_state.value) if stage_state else 0
    
    if next_stage != current_stage + 1:
        # Allow force advance? Or error.
        # Strict for now:
         raise HTTPException(status_code=400, detail=f"No se puede avanzar a la etapa {next_stage} desde la etapa {current_stage}")

    # Logica de calculo de diferencias (Copied from advance_inventory_stage)
    prev_stage = next_stage - 1
    stmt = select(StockCount.item_code, func.sum(StockCount.counted_qty).label('total_counted')).\
        join(CountSession, StockCount.session_id == CountSession.id).\
        where(CountSession.inventory_stage == prev_stage, CountSession.country_code == country, StockCount.country_code == country).\
        group_by(StockCount.item_code)
    
    result = await db.execute(stmt)
    counted_items = result.all()
    
    await db.execute(delete(RecountList).where(RecountList.stage_to_count == next_stage, RecountList.country_code == country))

    items_for_recount = []
    country_master_qty = master_qty_map.get(country, {})
    for item in counted_items:
        item_code = item.item_code
        total_counted = item.total_counted
        system_qty = country_master_qty.get(item_code)
        system_qty = int(system_qty) if system_qty is not None else 0

        if total_counted != system_qty:
            items_for_recount.append({"item_code": item_code, "stage_to_count": next_stage, "country_code": country})

    if items_for_recount:
        await db.execute(insert(RecountList), items_for_recount)
    
    stage_state.value = str(next_stage)
    await db.commit()
    from fastapi.responses import JSONResponse
    return JSONResponse(content={"message": f"Avanzado a Etapa {next_stage} en {country}", "stage": next_stage})

@router.post('/api/admin/inventory/finalize')
async def finalize_inventory_api(request: Request, user: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """API: Finaliza inventario."""
    country = get_current_country(request) or "CL"
    result = await db.execute(select(AppState).where(AppState.key == 'current_inventory_stage', AppState.country_code == country))
    stage_state = result.scalar_one_or_none()
    if stage_state:
        stage_state.value = '0' 
        await db.commit()
    from fastapi.responses import JSONResponse
    return JSONResponse(content={"message": f"Inventario en {country} finalizado correctamente", "stage": 0})


# ===== RUTAS DE MANAGE COUNTS =====

@router.get('/manage_counts', response_class=HTMLResponse, name='manage_counts_page')
async def manage_counts_page(request: Request, username: str = Depends(permission_required("inventory")), db: AsyncSession = Depends(get_db)):
    """Página de gestión de conteos."""
    if not isinstance(username, str):
        return username
    
    country = get_current_country(request) or "CL"
    counts = await db_counts.load_all_counts_db_async(db, country_code=country)
    
    return templates.TemplateResponse('manage_counts.html', {"request": request, "counts": counts})
