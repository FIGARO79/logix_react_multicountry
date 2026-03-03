from fastapi import APIRouter, Request, Form, Depends, HTTPException, status, File, UploadFile, Response, BackgroundTasks
from fastapi.responses import JSONResponse, RedirectResponse, HTMLResponse
import pandas as pd
import os
import shutil
import json
import datetime
from urllib.parse import urlencode
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import delete
from app.core.db import get_db
from app.models.sql_models import Log

# Importaciones relativas desde la estructura del proyecto
from app.core.config import (
    ITEM_MASTER_CSV_PATH,
    GRN_CSV_FILE_PATH,
    PICKING_CSV_PATH,
    GRN_COLUMN_NAME_IN_CSV,
    ADMIN_PASSWORD
)
from app.services.csv_handler import load_csv_data, get_country_csv_path, DATABASE_FOLDER
from app.utils.country import get_current_country
from app.utils.auth import login_required
from app.core.templates import templates

router = APIRouter(
    prefix="",
    tags=["update"]
)

# --- Endpoint para la página de actualización (GET) ---
@router.get('/update', response_class=HTMLResponse)
async def update_files_get(request: Request, username: str = Depends(login_required)):
    if not isinstance(username, str):
        return username  # Devuelve la redirección si el login falla
    
    return templates.TemplateResponse("update.html", {
        "request": request,
        "error": request.query_params.get('error'),
        "message": request.query_params.get('message')
    })

# --- Endpoint para subir y procesar los archivos (POST) ---
@router.post('/api/update', response_class=JSONResponse)
async def update_files_post(
    request: Request,
    background_tasks: BackgroundTasks,
    item_master: UploadFile = File(None),
    grn_file: UploadFile = File(None),
    picking_file: UploadFile = File(None),
    update_option_280: str = Form(None),
    selected_grns_280: str = Form(None),
    username: str = Depends(login_required)
):
    if not isinstance(username, str):
        return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"error": "Unauthorized"})

    files_uploaded = False
    message = ""
    error = ""

    country = get_current_country(request) or "MX"
    
    # Manejo del maestro de items
    if item_master and item_master.filename:
        target_path = get_country_csv_path(DATABASE_FOLDER, 'AURRSGLBD0250.csv', country)
        with open(target_path, "wb") as buffer:
            shutil.copyfileobj(item_master.file, buffer)
        message += f'Archivo "{item_master.filename}" actualizado (Maestro) para {country}. '
        files_uploaded = True

    # Manejo del archivo GRN (280)
    if grn_file and grn_file.filename:
        try:
            target_path_grn = get_country_csv_path(DATABASE_FOLDER, 'AURRSGLBD0280.csv', country)
            new_data_df = pd.read_csv(grn_file.file, dtype=str)
            
            if selected_grns_280:
                try:
                    selected_list = json.loads(selected_grns_280)
                    if selected_list:
                        original_count = len(new_data_df)
                        new_data_df = new_data_df[new_data_df[GRN_COLUMN_NAME_IN_CSV].isin(selected_list)]
                        message += f"Filtrado: {len(new_data_df)} registros de {original_count}. "
                except json.JSONDecodeError:
                    pass

            if update_option_280 == 'combine':
                if os.path.exists(target_path_grn):
                    existing_data_df = pd.read_csv(target_path_grn, dtype=str)
                    
                    # Obtener las GRNs que vienen en el archivo nuevo
                    new_grns = new_data_df[GRN_COLUMN_NAME_IN_CSV].unique()
                    
                    # Eliminar del archivo existente todas las líneas de las GRNs que vienen en el nuevo archivo
                    existing_data_df = existing_data_df[~existing_data_df[GRN_COLUMN_NAME_IN_CSV].isin(new_grns)]
                    
                    # Combinar
                    combined_df = pd.concat([existing_data_df, new_data_df], ignore_index=True)
                else:
                    combined_df = new_data_df

                combined_df.to_csv(target_path_grn, index=False)
                message += f'Archivo "{grn_file.filename}" combinado para {country}. '
            
            else:  # 'replace'
                new_data_df.to_csv(target_path_grn, index=False)
                message += f'Archivo "{grn_file.filename}" reemplazado para {country}. '
            
            files_uploaded = True
        except Exception as e:
            import traceback
            print(f"ERROR procesando archivo GRN: {str(e)}")
            print(traceback.format_exc())
            error += f'Error procesando archivo GRN: {str(e)}. '

    # Manejo del archivo de picking (240)
    if picking_file and picking_file.filename:
        target_path_picking = get_country_csv_path(DATABASE_FOLDER, 'AURRSGLBD0240.csv', country)
        with open(target_path_picking, "wb") as buffer:
            shutil.copyfileobj(picking_file.file, buffer)
        message += f'Archivo "{picking_file.filename}" actualizado (Picking) para {country}. '
        files_uploaded = True

    if files_uploaded:
        # Procesar en segundo plano
        background_tasks.add_task(load_csv_data, country_code=country)
        message += " Procesamiento de datos iniciado en segundo plano."

    if not files_uploaded and not error:
        error = "No seleccionaste ningún archivo para subir."

    if error:
        return JSONResponse(status_code=status.HTTP_400_BAD_REQUEST, content={"error": error})
    
    return JSONResponse(content={"message": message})


# --- Endpoint para previsualizar las GRNs de un archivo ---
@router.post("/api/preview_grn_file")
async def preview_grn_file(file: UploadFile = File(...), username: str = Depends(login_required)):
    try:
        contents = await file.read()
        df = pd.read_csv(BytesIO(contents), dtype=str, keep_default_na=True)
        
        if GRN_COLUMN_NAME_IN_CSV not in df.columns:
            return JSONResponse(
                status_code=400, 
                content={"error": f"No se encontró la columna {GRN_COLUMN_NAME_IN_CSV} en el archivo."}
            )
        
        grns = df[GRN_COLUMN_NAME_IN_CSV].dropna().unique().tolist()
        grns.sort()
        
        return JSONResponse(content={"grns": grns})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


# --- Endpoint para la "Zona de Peligro" de limpiar la BD ---
@router.post('/api/clear_database')
async def clear_database_api(request: Request, password: str = Form(...), db: AsyncSession = Depends(get_db)):
    """API: Limpia la base de datos de logs (Zona de Peligro)."""
    if password != ADMIN_PASSWORD:
        return JSONResponse(status_code=401, content={"error": "Contraseña incorrecta"})
    
    country = get_current_country(request) or "MX"
    try:
        await db.execute(delete(Log).where(Log.country_code == country))
        await db.commit()
        return JSONResponse(content={"message": f"Base de datos de logs para {country} limpiada correctamente"})
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

@router.post('/clear_database')
async def clear_database(request: Request, password: str = Form(...), db: AsyncSession = Depends(get_db)):
    country = get_current_country(request) or "MX"
    # La URL de redirección debe ser construida correctamente
    redirect_url = request.url_for('update_files_get')

    if password != ADMIN_PASSWORD:
        query_params = urlencode({'error': 'Contraseña incorrecta'})
        return RedirectResponse(url=f'{redirect_url}?{query_params}', status_code=status.HTTP_302_FOUND)
    
    try:
        await db.execute(delete(Log).where(Log.country_code == country))
        await db.commit()
        
        query_params = urlencode({'message': f'Base de datos de logs para {country} limpiada'})
        return RedirectResponse(url=f'{redirect_url}?{query_params}', status_code=status.HTTP_302_FOUND)
    
    except Exception as e:
        query_params = urlencode({'error': str(e)})
        return RedirectResponse(url=f'{redirect_url}?{query_params}', status_code=status.HTTP_302_FOUND)


# --- Endpoint para descargar TODO el log (Backup) ---
@router.post('/api/export_all_log')
async def export_all_log_api(request: Request, password: str = Form(...), db: AsyncSession = Depends(get_db)):
    """API: Exporta TODOS los registros (activos y archivados)."""
    if password != ADMIN_PASSWORD:
         return JSONResponse(status_code=401, content={"error": "Contraseña incorrecta"})

    country = get_current_country(request) or "MX"
    try:
        from app.services import db_logs
        from openpyxl.utils import get_column_letter

        # Reuse logic? Copied for safety and speed.
        logs_data = await db_logs.load_all_logs_db_async(db, country_code=country)
        
        if not logs_data:
             return JSONResponse(status_code=404, content={"error": f"No hay datos para exportar backup en {country}"})

        df = pd.DataFrame(logs_data)
        try:
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            if df['timestamp'].dt.tz is not None:
                 colombia_tz = datetime.timezone(datetime.timedelta(hours=-5))
                 df['timestamp'] = df['timestamp'].dt.tz_convert(colombia_tz).dt.tz_localize(None)
        except Exception:
            pass
        
        if 'archived_at' in df.columns:
             df['archived_at'] = df['archived_at'].fillna('Activo')

        df_export = df.rename(columns={
            'timestamp': 'Timestamp', 'importReference': 'Import Reference', 'waybill': 'Waybill',
            'itemCode': 'Item Code', 'itemDescription': 'Item Description',
            'binLocation': 'Bin Location (Original)', 'relocatedBin': 'Relocated Bin (New)',
            'qtyReceived': 'Qty. Received', 'qtyGrn': 'Qty. Expected (Total)', 'difference': 'Difference',
            'archived_at': 'Estado / Fecha Archivo'
        })
        
        cols = ['Timestamp', 'Import Reference', 'Waybill', 'Item Code', 'Item Description', 
                'Bin Location (Original)', 'Relocated Bin (New)', 'Qty. Received', 
                'Qty. Expected (Total)', 'Difference', 'Estado / Fecha Archivo']
        cols = [c for c in cols if c in df_export.columns]
        df_export = df_export[cols]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='BackupCompleto')
            worksheet = writer.sheets['BackupCompleto']
            for i, col_name in enumerate(df_export.columns):
                column_letter = get_column_letter(i + 1)
                max_len = max(df_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[column_letter].width = max_len

        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"BACKUP_FULL_LOG_{timestamp_str}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        return JSONResponse(status_code=500, content={"error": f"Error generando backup: {str(e)}"})

# Old endpoint kept for legacy safety
@router.post('/export_all_log')
async def export_all_log(request: Request, password: str = Form(...), db: AsyncSession = Depends(get_db)):
    """Exporta TODOS los registros (activos y archivados) a Excel como backup."""
    redirect_url = request.url_for('update_files_get')

    if password != ADMIN_PASSWORD:
        query_params = urlencode({'error': 'Contraseña incorrecta'})
        return RedirectResponse(url=f'{redirect_url}?{query_params}', status_code=status.HTTP_302_FOUND)

    try:
        from app.services import db_logs
        from openpyxl.utils import get_column_letter

        # Cargar TODOS los logs
        logs_data = await db_logs.load_all_logs_db_async(db)
        
        if not logs_data:
             query_params = urlencode({'error': 'No hay datos para exportar backup'})
             return RedirectResponse(url=f'{redirect_url}?{query_params}', status_code=status.HTTP_302_FOUND)

        df = pd.DataFrame(logs_data)

        # Procesar timestamp
        try:
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            if df['timestamp'].dt.tz is not None:
                 colombia_tz = datetime.timezone(datetime.timedelta(hours=-5))
                 df['timestamp'] = df['timestamp'].dt.tz_convert(colombia_tz).dt.tz_localize(None)
        except Exception:
            pass
        
        # Procesar archived_at para formato limpio
        if 'archived_at' in df.columns:
             df['archived_at'] = df['archived_at'].fillna('Activo')

        df_export = df.rename(columns={
            'timestamp': 'Timestamp', 'importReference': 'Import Reference', 'waybill': 'Waybill',
            'itemCode': 'Item Code', 'itemDescription': 'Item Description',
            'binLocation': 'Bin Location (Original)', 'relocatedBin': 'Relocated Bin (New)',
            'qtyReceived': 'Qty. Received', 'qtyGrn': 'Qty. Expected (Total)', 'difference': 'Difference',
            'archived_at': 'Estado / Fecha Archivo'
        })

        # Columnas a exportar (asegurando orden)
        cols = ['Timestamp', 'Import Reference', 'Waybill', 'Item Code', 'Item Description', 
                'Bin Location (Original)', 'Relocated Bin (New)', 'Qty. Received', 
                'Qty. Expected (Total)', 'Difference', 'Estado / Fecha Archivo']
        
        # Filtrar solo columnas existentes (por si acaso)
        cols = [c for c in cols if c in df_export.columns]
        df_export = df_export[cols]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='BackupCompleto')
            worksheet = writer.sheets['BackupCompleto']
            for i, col_name in enumerate(df_export.columns):
                column_letter = get_column_letter(i + 1)
                max_len = max(df_export[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[column_letter].width = max_len

        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"BACKUP_FULL_LOG_{timestamp_str}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        query_params = urlencode({'error': f'Error generando backup: {str(e)}'})
        return RedirectResponse(url=f'{redirect_url}?{query_params}', status_code=status.HTTP_302_FOUND)
