"""
Router para endpoints de conteo.
"""
import datetime
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, status, Request
from fastapi.responses import JSONResponse, Response
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy import select, func, distinct
from app.core.db import get_db
import numpy as np

from app.models.schemas import StockCount, Count
from app.models.sql_models import CountSession, RecountList, StockCount as StockCountModel, CycleCountRecording
from app.services import db_counts, csv_handler
from app.services.csv_handler import master_qty_map, get_locations_with_stock_count # Importar el mapa de memoria y helper
from app.utils.auth import login_required, api_login_required
from app.utils.country import get_current_country
from app.core.config import ASYNC_DB_URL

# --- Inicialización de elementos compartidos ---
router = APIRouter(prefix="/api", tags=["counts"])
async_engine = create_async_engine(ASYNC_DB_URL)


@router.get('/get_item_for_counting/{item_code}')
async def get_item_for_counting(request: Request, item_code: str, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    country = get_current_country(request) or "MX"
    current_stage = 1 # Default
    
    # Obtener sesión activa ORM para el país
    result = await db.execute(
        select(CountSession)
        .where(CountSession.user_username == username, CountSession.status == 'in_progress', CountSession.country_code == country)
        .order_by(CountSession.start_time.desc())
        .limit(1)
    )
    active_session = result.scalar_one_or_none()
    
    if not active_session:
        raise HTTPException(status_code=403, detail="No tienes una sesión de conteo activa. Inicia una nueva sesión.")
        
    current_stage = active_session.inventory_stage

    if current_stage > 1:
        result_recount = await db.execute(
            select(RecountList).where(RecountList.item_code == item_code, RecountList.stage_to_count == current_stage, RecountList.country_code == country)
        )
        if not result_recount.scalar_one_or_none():
            raise HTTPException(status_code=404, detail=f"Item no requerido. Este item no está en la lista de reconteo para la Etapa {current_stage}.")

    details = await csv_handler.get_item_details_from_master_csv(item_code, country_code=country)
    if details:
        return JSONResponse(content={
            'item_code': details.get('Item_Code'),
            'description': details.get('Item_Description'),
            'bin_location': details.get('Bin_1')
        })
    else:
        if current_stage == 1:
            return JSONResponse(content={
                'item_code': item_code,
                'description': 'ITEM NO ENCONTRADO',
                'bin_location': 'N/A'
            })
        else:
            raise HTTPException(status_code=404, detail="Artículo no encontrado en el maestro de items.")


@router.get('/counts/all')
async def get_all_counts(request: Request, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Obtiene todos los registros de conteo con información enriquecida para el país."""
    country = get_current_country(request) or "MX"
    try:
        all_counts = await db_counts.load_all_counts_db_async(db, country_code=country)
        
        # Obtener detalles de sesiones del país
        session_ids = list({c.get('session_id') for c in all_counts if c.get('session_id') is not None})
        session_map = {}
        
        if session_ids:
            result = await db.execute(select(CountSession).where(CountSession.id.in_(session_ids), CountSession.country_code == country))
            sessions = result.scalars().all()
            session_map = {s.id: {'user': s.user_username, 'stage': s.inventory_stage} for s in sessions}

        # Asegurar que el cache esté cargado para el país
        if country not in master_qty_map:
            await csv_handler.reload_cache_if_needed(country)
        
        country_master_map = master_qty_map.get(country, {})

        enriched_counts = []
        for count in all_counts:
            item_code = count.get('item_code')
            system_qty_raw = country_master_map.get(item_code)
            system_qty = int(float(system_qty_raw)) if system_qty_raw is not None else None
            counted_qty = int(count.get('counted_qty', 0))
            difference = (counted_qty - system_qty) if system_qty is not None else None
            
            session_info = session_map.get(count.get('session_id'), {})
            
            enriched_counts.append({
                'id': count.get('id'),
                'session_id': count.get('session_id'),
                'inventory_stage': session_info.get('stage'),
                'username': count.get('username') or session_info.get('user'),
                'timestamp': str(count.get('timestamp', '')).split('.')[0] if count.get('timestamp') else '',
                'item_code': item_code,
                'item_description': count.get('item_description'),
                'counted_location': count.get('counted_location'),
                'counted_qty': counted_qty,
                'system_qty': system_qty,
                'difference': difference,
                'bin_location_system': count.get('bin_location_system')
            })
        
        return JSONResponse(content=enriched_counts)
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {e}")


@router.post('/counts')
async def add_count(data: Count, username: str = Depends(api_login_required)):
    """Añade un conteo básico."""
    return JSONResponse({'message': 'Endpoint de conteo no implementado en esta versión.'}, status_code=501)


@router.post('/save_count')
async def save_count(request: Request, data: StockCount, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Guarda un conteo de stock con sesión para el país."""
    country = get_current_country(request) or "MX"
    # Lógica existente...
    item_details = await csv_handler.get_item_details_from_master_csv(data.item_code, country_code=country)
    description = item_details.get('Item_Description', '') if item_details else data.description or ''
    bin_location_system = item_details.get('Bin_1', '') if item_details else data.bin_location_system or ''

    count_id = await db_counts.save_stock_count(
        db,
        session_id=data.session_id,
        item_code=data.item_code,
        counted_qty=data.counted_qty,
        counted_location=data.counted_location,
        description=description,
        bin_location_system=bin_location_system,
        username=username,
        country_code=country
    )
    if count_id:
        return JSONResponse({'message': 'Conteo guardado con éxito.', 'countId': count_id}, status_code=201)
    raise HTTPException(status_code=500, detail="Error al guardar el conteo.")


@router.get('/counts/differences')
async def get_count_differences(request: Request, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Obtiene diferencias para el país."""
    country = get_current_country(request) or "MX"
    try:
        all_counts = await db_counts.load_all_counts_db_async(db, country_code=country)
        
        # Obtener detalles de sesiones
        session_ids = list({c.get('session_id') for c in all_counts if c.get('session_id') is not None})
        session_map = {}
        
        if session_ids:
            result = await db.execute(select(CountSession).where(CountSession.id.in_(session_ids), CountSession.country_code == country))
            sessions = result.scalars().all()
            session_map = {s.id: {'user': s.user_username, 'stage': s.inventory_stage} for s in sessions}

        if country not in master_qty_map:
            await csv_handler.reload_cache_if_needed(country)
        country_master_map = master_qty_map.get(country, {})

        differences_list = []
        
        for count in all_counts:
            item_code = count.get('item_code')
            system_qty_raw = country_master_map.get(item_code)
            system_qty = int(float(system_qty_raw)) if system_qty_raw is not None else 0
            counted_qty = int(count.get('counted_qty', 0))
            difference = counted_qty - system_qty
            
            # Calcular porcentaje de varianza
            percentage_variance = 0
            if system_qty > 0:
                percentage_variance = round((difference / system_qty) * 100, 2)
            
            session_info = session_map.get(count.get('session_id'), {})
            
            differences_list.append({
                'count_id': count.get('id'),
                'item_code': item_code,
                'description': count.get('item_description', 'N/A'),
                'location': count.get('counted_location', 'N/A'),
                'system_qty': system_qty,
                'counted_qty': counted_qty,
                'difference': difference,
                'percentage_variance': percentage_variance,
                'date': str(count.get('timestamp', '')).split('.')[0] if count.get('timestamp') else '',
                'username': count.get('username') or session_info.get('user', 'N/A')
            })
        
        # Ordenar por diferencia (mayor primero)
        differences_list.sort(key=lambda x: abs(x['difference']), reverse=True)
        
        return JSONResponse(content={
            'total': len(differences_list),
            'items': differences_list
        })
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {e}")


@router.put("/counts/{count_id}")
async def update_stock_count(request: Request, count_id: int, data: dict, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Actualiza un conteo filtrado por país."""
    country = get_current_country(request) or "MX"
    try:
        # Obtener el registro actual validando país
        result = await db.execute(
            select(StockCountModel).where(StockCountModel.id == count_id, StockCountModel.country_code == country)
        )
        count = result.scalar_one_or_none()
        
        if not count:
            raise HTTPException(status_code=404, detail=f"Conteo con ID {count_id} no encontrado.")
        
        # Actualizar la cantidad
        count.counted_qty = data.get('counted_qty', count.counted_qty)
        
        await db.commit()
        
        return JSONResponse(content={
            'message': 'Cantidad actualizada exitosamente',
            'count_id': count_id
        }, status_code=200)
    
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al actualizar el conteo: {e}")


@router.delete("/counts/{count_id}", status_code=status.HTTP_200_OK)
async def delete_stock_count(request: Request, count_id: int, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Elimina un conteo de stock validando país."""
    country = get_current_country(request) or "MX"
    success = await db_counts.delete_stock_count(db, count_id, country_code=country)
    if success:
        return JSONResponse({'message': f'Conteo {count_id} eliminado con éxito.'})
    raise HTTPException(status_code=404, detail=f"Conteo con ID {count_id} no encontrado.")


@router.get('/export_counts')
async def export_counts(request: Request, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Exporta conteos del país a Excel."""
    country = get_current_country(request) or "MX"
    all_counts = await db_counts.load_all_counts_db_async(db, country_code=country)
    
    # Obtener detalles de sesiones
    session_ids = list({c.get('session_id') for c in all_counts if c.get('session_id') is not None})
    session_map = {}
    
    if session_ids:
        result = await db.execute(select(CountSession).where(CountSession.id.in_(session_ids), CountSession.country_code == country))
        sessions = result.scalars().all()
        session_map = {s.id: {'user': s.user_username, 'stage': s.inventory_stage} for s in sessions}

    if country not in master_qty_map:
        await csv_handler.reload_cache_if_needed(country)
    country_master_map = master_qty_map.get(country, {})

    enriched_rows = []
    for count in all_counts:
        item_code = count.get('item_code')
        system_qty_raw = country_master_map.get(item_code)
        system_qty = int(float(system_qty_raw)) if system_qty_raw is not None else None
        counted_qty = int(count.get('counted_qty', 0))
        difference = (counted_qty - system_qty) if system_qty is not None else None
        
        session_info = session_map.get(count.get('session_id'), {})
        enriched = {
            'id': count.get('id'),
            'session_id': count.get('session_id'),
            'inventory_stage': session_info.get('stage'),
            'username': count.get('username') or session_info.get('user'),
            'timestamp': count.get('timestamp'),
            'item_code': item_code,
            'item_description': count.get('item_description'),
            'counted_location': count.get('counted_location'),
            'counted_qty': counted_qty,
            'system_qty': system_qty,
            'difference': difference,
            'bin_location_system': count.get('bin_location_system')
        }
        enriched_rows.append(enriched)

    df = pd.DataFrame(enriched_rows)
    df = df[['id', 'session_id', 'inventory_stage', 'username', 'timestamp', 'item_code', 'item_description', 'counted_location', 'counted_qty', 'system_qty', 'difference', 'bin_location_system']]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Conteos')
        worksheet = writer.sheets['Conteos']
        for i, col_name in enumerate(df.columns):
            column_letter = get_column_letter(i + 1)
            max_len = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
            worksheet.column_dimensions[column_letter].width = max_len

    output.seek(0)
    timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"conteos_export_{timestamp_str}.xlsx"
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get('/counts/stats')
async def get_count_stats(request: Request, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Estadísticas del país."""
    country = get_current_country(request) or "MX"
    try:
        # Asegurar cache
        if country not in master_qty_map:
            await csv_handler.reload_cache_if_needed(country)
        country_master_map = master_qty_map.get(country, {})

        # 1. Total de ubicaciones contadas
        result_locs = await db.execute(select(func.count(distinct(StockCountModel.counted_location))).where(StockCountModel.country_code == country))
        counted_locations = result_locs.scalar() or 0
        
        # 2. Total de items contados
        result_items = await db.execute(select(func.count(distinct(StockCountModel.item_code))).where(StockCountModel.country_code == country))
        total_items_counted = result_items.scalar() or 0
        
        # 3. Total de items con stock
        total_items_with_stock = sum(1 for qty in country_master_map.values() if qty is not None and qty > 0)
        
        # 4. Items con diferencias
        stmt = select(
            StockCountModel.item_code,
            func.sum(StockCountModel.counted_qty).label('total_counted')
        ).where(StockCountModel.country_code == country).group_by(StockCountModel.item_code)
        
        result_agg = await db.execute(stmt)
        all_counted_items = result_agg.all()
        counted_qty_map = {item.item_code: item.total_counted for item in all_counted_items}

        items_with_differences = 0
        items_with_positive_differences = 0
        items_with_negative_differences = 0

        for item_code, total_counted in counted_qty_map.items():
            system_qty_raw = country_master_map.get(item_code)
            system_qty = 0
            if system_qty_raw is not None:
                try:
                    system_qty = int(float(system_qty_raw))
                except (ValueError, TypeError):
                    system_qty = 0
            
            if total_counted != system_qty:
                items_with_differences += 1
                if total_counted > system_qty:
                    items_with_positive_differences += 1
                else:
                    items_with_negative_differences += 1

        # 5. Total de ubicaciones con stock para el país
        total_locations_with_stock = await get_locations_with_stock_count(country_code=country)

        return JSONResponse(content={
            "total_items_with_stock": total_items_with_stock,
            "counted_locations": counted_locations,
            "total_items_counted": total_items_counted,
            "items_with_differences": items_with_differences,
            "items_with_positive_differences": items_with_positive_differences,
            "items_with_negative_differences": items_with_negative_differences,
            "total_locations_with_stock": total_locations_with_stock
        })
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {e}")


@router.get('/counts/debug/master_qty_map')
async def debug_master_qty_map(request: Request, username: str = Depends(api_login_required)):
    """Debug map del país."""
    country = get_current_country(request) or "MX"
    country_map = master_qty_map.get(country, {})
    total_items = len(country_map)
    items_with_stock = sum(1 for qty in country_map.values() if qty is not None and qty > 0)
    sample_items = dict(list(country_map.items())[:10])
    
    return JSONResponse(content={
        "country": country,
        "total_items_in_map": total_items,
        "items_with_stock": items_with_stock,
        "sample_items": sample_items,
        "map_is_empty": len(country_map) == 0
    })


@router.get('/debug/last_counts')
async def debug_last_counts(request: Request, limit: int = 20, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Diágnostico por país."""
    country = get_current_country(request) or "MX"
    all_counts = await db_counts.load_all_counts_db_async(db, country_code=country)
    return JSONResponse(content=all_counts[:int(limit)])


@router.get('/counts/recordings')
async def get_cycle_count_recordings(request: Request, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Historial del país."""
    country = get_current_country(request) or "MX"
    from app.models.sql_models import MasterItem
    
    try:
        result = await db.execute(
            select(CycleCountRecording).where(CycleCountRecording.country_code == country).order_by(CycleCountRecording.executed_date.desc())
        )
        recordings = result.scalars().all()
        
        if not recordings:
            return JSONResponse(content=[])
        
        # OPTIMIZACIÓN: Batch query para todos los item codes de una vez
        item_codes = list({rec.item_code for rec in recordings})
        
        # Consultar todos los items necesarios del país en una sola query
        result_items = await db.execute(
            select(MasterItem).where(MasterItem.item_code.in_(item_codes), MasterItem.country_code == country)
        )
        master_items = result_items.scalars().all()
        
        # Crear un mapa para lookup rápido
        master_map = {item.item_code: item for item in master_items}
        
        enriched_data = []
        for rec in recordings:
            # Buscar detalles en el mapa (O(1) lookup)
            master_item = master_map.get(rec.item_code)
            
            # Valores por defecto
            cost = 0.0
            stockroom = "N/A"
            item_type = ""
            item_class = ""
            item_group = ""
            sic_company = ""
            sic_stockroom = ""
            weight = ""
            
            if master_item:
                # Leer desde la tabla master_items
                stockroom = master_item.stockroom or "N/A"
                item_type = master_item.item_type or ""
                item_class = master_item.item_class or ""
                item_group = master_item.item_group_major or ""
                sic_company = master_item.sic_code_company or ""
                sic_stockroom = master_item.sic_code_stockroom or ""
                weight = str(master_item.weight_per_unit) if master_item.weight_per_unit else ""
                
                try:
                    cost = float(master_item.cost_per_unit) if master_item.cost_per_unit else 0.0
                except (ValueError, TypeError):
                    cost = 0.0
            
            value_diff = rec.difference * cost
            count_value = rec.physical_qty * cost
            
            enriched_data.append({
                'id': rec.id,
                'executed_date': rec.executed_date,
                'username': rec.username,
                'item_code': rec.item_code,
                'description': rec.item_description,
                'stockroom': stockroom,
                'item_type': item_type,
                'item_class': item_class,
                'item_group': item_group,
                'sic_company': sic_company,
                'sic_stockroom': sic_stockroom,
                'weight': weight,
                'abc_code': rec.abc_code,
                'bin_location': rec.bin_location,
                'physical_qty': rec.physical_qty,
                'system_qty': rec.system_qty,
                'difference': rec.difference,
                'cost': cost,
                'value_diff': value_diff,
                'count_value': count_value
            })
            
        return JSONResponse(content=enriched_data)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo historial: {e}")


@router.get('/counts/export_recordings')
async def export_cycle_count_recordings(request: Request, username: str = Depends(api_login_required), db: AsyncSession = Depends(get_db)):
    """Exporta historial por país."""
    country = get_current_country(request) or "MX"
    try:
        result = await db.execute(
            select(CycleCountRecording).where(CycleCountRecording.country_code == country).order_by(CycleCountRecording.executed_date.desc())
        )
        recordings = result.scalars().all()
        
        enriched_data = []
        for rec in recordings:
            details = await csv_handler.get_item_details_from_master_csv(rec.item_code, country_code=country)
            cost = 0.0
            stockroom = "N/A"
            
            if details:
                stockroom = details.get('Stockroom', 'N/A')
                try:
                    cost_str = details.get('Cost_per_Unit', '0').replace(',', '')
                    cost = float(cost_str)
                except (ValueError, AttributeError):
                    cost = 0.0
            
            value_diff = rec.difference * cost
            count_value = rec.physical_qty * cost
            
            enriched_data.append({
                'ID': rec.id,
                'Stockroom': stockroom,
                'Item Code': rec.item_code,
                'Description': rec.item_description,
                'Type': details.get('Item_Type', '') if details else '',
                'Class': details.get('Item_Class', '') if details else '',
                'Group': details.get('Item_Group_Major', '') if details else '',
                'SICs': details.get('SIC_Code_Company', '') if details else '',
                'Weight': details.get('Weight_per_Unit', '') if details else '',
                'ABC': rec.abc_code,
                'Bin': rec.bin_location,
                'Sys Stock': rec.system_qty,
                'Counted': rec.physical_qty,
                'Diff': rec.difference,
                'Value Diff': value_diff,
                'Item Cost': cost,
                'Count Value': count_value,
                'Date': rec.executed_date,
                'User': rec.username
            })
            
        df = pd.DataFrame(enriched_data)
        
        # Reorder columns as per UI
        columns_order = [
            'ID', 'Stockroom', 'Item Code', 'Description', 'Type', 'Class', 'Group', 
            'SICs', 'Weight', 'ABC', 'Bin', 'Sys Stock', 'Counted', 'Diff', 
            'Value Diff', 'Item Cost', 'Count Value', 'Date', 'User'
        ]
        # Ensure only existing columns are selected (in case some are missing keys)
        df = df[[c for c in columns_order if c in df.columns]]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='History')
            worksheet = writer.sheets['History']
            for i, col_name in enumerate(df.columns):
                column_letter = get_column_letter(i + 1)
                max_len = max(df[col_name].astype(str).map(len).max(), len(col_name)) + 2
                worksheet.column_dimensions[column_letter].width = max_len

        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cycle_count_history_{timestamp_str}.xlsx"
        
        return Response(
            content=output.getvalue(), 
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exportando historial: {e}")
